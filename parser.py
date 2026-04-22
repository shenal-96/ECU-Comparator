"""Parse ECU parameter .XLS files into structured dicts."""

from openpyxl import load_workbook
from pathlib import Path
from typing import Dict, Any


def parse_file(filepath: str) -> Dict[str, Any]:
    """Load and parse a single XLS/XLSX file.

    Returns dict with 'label' and 'sheets' (containing Parameter, Val_2D, Val_3D).
    """
    wb = load_workbook(filepath, data_only=True)

    label = Path(filepath).stem
    sheets = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if sheet_name == "Parameter":
            sheets["Parameter"] = parse_parameter(sheet)
        elif sheet_name == "Val_2D":
            sheets["Val_2D"] = parse_val_2d(sheet)
        elif sheet_name == "Val_3D":
            sheets["Val_3D"] = parse_val_3d(sheet)

    return {"label": label, "sheets": sheets}


def parse_parameter(sheet) -> Dict[str, Dict[str, Any]]:
    """Parse Parameter sheet: Nr -> {name, value, unit}."""
    result = {}

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        try:
            nr = row[0]
            if not nr:
                continue

            name = row[1] if len(row) > 1 else ""
            value = row[2] if len(row) > 2 else None
            unit = row[3] if len(row) > 3 else ""

            try:
                value = float(value) if value else None
            except (ValueError, TypeError):
                value = None

            nr_key = str(int(nr) if isinstance(nr, float) else nr)
            result[nr_key] = {
                "name": str(name) if name else "",
                "value": value,
                "unit": str(unit) if unit else "",
            }
        except Exception:
            continue

    return result


def parse_val_2d(sheet) -> Dict[str, Dict[str, Any]]:
    """Parse Val_2D sheet. Each curve = 4 rows: [id, x-axis, y-axis, blank]."""
    result = {}

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    row_idx = 0

    while row_idx < len(rows):
        try:
            nr = rows[row_idx][0]
            if not nr:
                row_idx += 1
                continue

            name = rows[row_idx][1] if len(rows[row_idx]) > 1 else ""
            nr_key = str(int(nr) if isinstance(nr, float) else nr)

            if row_idx + 2 >= len(rows):
                row_idx += 4
                continue

            x_row = rows[row_idx + 1]
            y_row = rows[row_idx + 2]

            x_values = []
            y_values = []

            for col in range(2, len(x_row)):
                x_val = x_row[col]
                y_val = y_row[col]

                try:
                    x_values.append(float(x_val) if x_val else 0)
                    y_values.append(float(y_val) if y_val else 0)
                except (ValueError, TypeError):
                    pass

            result[nr_key] = {
                "name": str(name) if name else "",
                "x_values": x_values,
                "y_values": y_values,
            }

            row_idx += 4
        except Exception:
            row_idx += 4

    return result


def parse_val_3d(sheet) -> Dict[str, Dict[str, Any]]:
    """Parse Val_3D sheet. Each map = variable rows: [id+x, data rows..., blank]."""
    result = {}

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    row_idx = 0

    while row_idx < len(rows):
        try:
            nr = rows[row_idx][0]
            if not nr or nr == "":
                row_idx += 1
                continue

            # Skip rows without a numeric Nr (floating point number)
            if not isinstance(nr, (int, float)):
                row_idx += 1
                continue

            name = rows[row_idx][1] if len(rows[row_idx]) > 1 else ""
            nr_key = str(int(nr))

            # X-axis values start at column F (index 5) of the identifier row
            x_values = []
            for col in range(5, len(rows[row_idx])):
                x_val = rows[row_idx][col]
                if not x_val or x_val == "":
                    break
                try:
                    x_values.append(float(x_val))
                except (ValueError, TypeError):
                    break

            y_values = []
            grid = []
            data_row = row_idx + 1

            # Parse data rows until we hit a separator
            while data_row < len(rows):
                # Column E (index 4) contains y-axis breakpoints in data rows
                y_val = rows[data_row][4] if len(rows[data_row]) > 4 else None
                col_c = rows[data_row][2] if len(rows[data_row]) > 2 else None

                # Empty row = separator between maps
                if (not y_val or y_val == "") and (not col_c or col_c == ""):
                    break

                # Skip "rpm" header row
                if isinstance(col_c, str) and col_c.lower() == "rpm":
                    data_row += 1
                    continue

                # Parse y-axis value (must be numeric)
                try:
                    y_num = float(y_val) if y_val else None
                except (ValueError, TypeError):
                    data_row += 1
                    continue

                if y_num is None:
                    data_row += 1
                    continue

                y_values.append(y_num)

                # Grid values start at column F (index 5) matching x-axis column positions
                row_values = []
                for col in range(5, 5 + len(x_values)):
                    if col >= len(rows[data_row]):
                        break
                    cell_val = rows[data_row][col]
                    try:
                        row_values.append(float(cell_val) if cell_val != "" else 0)
                    except (ValueError, TypeError):
                        row_values.append(0)

                if row_values:
                    grid.append(row_values)

                data_row += 1

            result[nr_key] = {
                "name": str(name) if name else "",
                "x_values": x_values,
                "y_values": y_values,
                "grid": grid,
            }

            row_idx = data_row + 1
        except Exception:
            row_idx += 1

    return result
